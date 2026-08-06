#!/usr/bin/env python3
"""Генерация Excel-оценки проекта ОКО для коммерческого предложения."""
from __future__ import annotations

from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Установите openpyxl: pip install openpyxl") from exc

OUT = Path(__file__).resolve().parents[1] / "docs" / "OKO-оценка-трудозатрат.xlsx"
RATE_PER_HOUR = 4000

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
STAGE_FILL = PatternFill("solid", fgColor="D6E4F0")
STAGE_FONT = Font(bold=True, size=10)
TOTAL_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROLES = [
    "Руководитель проекта",
    "Аналитик / методолог",
    "Архитектор",
    "Backend-разработчик",
    "Frontend-разработчик",
    "Desktop-разработчик",
    "Тестировщик (QA)",
    "DevOps / инфраструктура",
]

STAGES = [
    {
        "id": "1",
        "name": "1. Обследование",
        "result": "Согласованные требования, границы проекта, критерии приёмки и перечень контрольных форм.",
        "works": [
            ("1.1", "Анализ бизнес-процессов заполнения и сдачи отчётности (Access)", "Карта текущих процессов и целевых сценариев.", [16, 40, 8, 0, 0, 0, 0, 0]),
            ("1.2", "Обследование эталонной базы z261.mdb: таблицы, формы, правила", "Описание структуры данных и правил legacy-системы.", [0, 24, 8, 16, 0, 0, 0, 0]),
            ("1.3", "Формирование требований к импортозамещению (ОС, СУБД, безопасность)", "Согласованный перечень технологических ограничений.", [8, 16, 8, 0, 0, 0, 0, 8]),
            ("1.4", "Согласование перечня форм и контрольных комплектов", "Эталонный набор форм для сдачи и тестов.", [8, 24, 0, 0, 0, 0, 8, 0]),
            ("1.5", "Формирование ФТ и НФТ, критериев приёмки", "Подписанные ФТ/НФТ и методика приёмки.", [16, 32, 8, 0, 0, 0, 0, 0]),
        ],
    },
    {
        "id": "2",
        "name": "2. Проектирование",
        "result": "Утверждённая архитектура, модель данных, API-контракты, UX-макеты и тест-план.",
        "works": [
            ("2.1", "Проектирование архитектуры решения и интеграций", "Архитектурная схема и протоколы взаимодействия.", [8, 8, 40, 8, 8, 8, 0, 8]),
            ("2.2", "Проектирование модели данных (PostgreSQL, миграции)", "ER-модель и стратегия миграций.", [0, 8, 24, 16, 0, 0, 0, 0]),
            ("2.3", "Проектирование UI/UX портала и десктопа", "Согласованные макеты ключевых экранов.", [8, 16, 8, 0, 40, 16, 0, 0]),
            ("2.4", "Проектирование расчётных бизнес-модулей (проверки, сальдо, расшифровки, пересчёт)", "Описание алгоритмов расчёта и проверки данных.", [0, 24, 16, 24, 8, 0, 0, 0]),
            ("2.5", "Проектирование SSO/LDAP, обмена комплектами, Excel/PDF", "Схемы интеграции с корпоративным контуром.", [8, 8, 16, 8, 0, 0, 0, 16]),
            ("2.6", "Разработка плана тестирования и приёмки", "Подробный план тестирования и актирования.", [8, 16, 0, 0, 0, 0, 24, 0]),
        ],
    },
    {
        "id": "3",
        "name": "3. Разработка",
        "result": "Готовая функциональная система: API, портал, офлайн-клиент, отчёты и миграция данных.",
        "works": [
            ("3.1", "API: авторизация, организации, периоды, экземпляры форм", "Рабочий API базового контура отчётности.", [0, 0, 8, 160, 0, 0, 0, 8]),
            ("3.2", "API: увязки, сальдо, расшифровки, агрегация, Excel-маппинг", "API методологии и расчётных функций.", [0, 8, 8, 200, 0, 0, 0, 0]),
            ("3.3", "API: пакеты, импорт/экспорт, версионирование методологии", "Управляемый обмен комплектами и версиями правил.", [0, 0, 0, 120, 0, 0, 0, 0]),
            ("3.4", "Портал: каталог форм, редактор, ячеечная модель, проверки", "Основной пользовательский контур заполнения форм.", [0, 0, 0, 0, 240, 0, 0, 0]),
            ("3.5", "Портал: редакторы методологии (формы, увязки, сальдо, расшифровки, Excel, агрегация)", "Админ-контур для методологов ЦО.", [0, 16, 0, 0, 280, 0, 0, 0]),
            ("3.6", "Портал: конструктор расшифровок (мастер, привязки, fixed/mixed, предпросмотр)", "Понятный конструктор расшифровок для бизнеса.", [0, 8, 0, 40, 120, 0, 0, 0]),
            ("3.7", "Портал: дашборд комплектов, обмен JSON, пользователи, инструкции", "Управление комплектами и ролями пользователей.", [0, 0, 0, 40, 120, 0, 0, 0]),
            ("3.8", "Разработка расчётных модулей для проверок, сальдо, пересчёта, расшифровок, агрегации, Excel/PDF", "Реализованы бизнес-алгоритмы, эквивалентные Access-логике.", [0, 8, 8, 120, 80, 0, 0, 0]),
            ("3.9", "Офлайн-клиент Tauri: SQLite, совместное заполнение, синхронизация", "Работа без интернета в сетевой папке подразделений.", [0, 0, 8, 40, 40, 320, 0, 0]),
            ("3.10", "Интеграция SSO / LDAP / Keycloak", "Единая корпоративная авторизация пользователей.", [0, 0, 8, 80, 16, 0, 0, 40]),
            ("3.11", "Экспорт Excel по шаблонам Минфина и PDF", "Регламентные выгрузки в требуемых форматах.", [0, 8, 0, 40, 40, 0, 0, 0]),
            ("3.12", "Миграция данных из MDB (скрипты, сверка)", "Перенос legacy-данных в новую платформу.", [0, 16, 0, 48, 0, 0, 24, 0]),
        ],
    },
    {
        "id": "4",
        "name": "4. Тестирование",
        "result": "Подтверждённое качество: функциональные, интеграционные, нагрузочные и security-проверки.",
        "works": [
            ("4.1", "Модульное и интеграционное тестирование (автотесты)", "Автоматизированные тесты ключевых модулей.", [0, 0, 0, 40, 24, 16, 80, 0]),
            ("4.2", "Функциональное тестирование по сценариям Access", "Сверка результата с эталонным поведением.", [0, 24, 0, 16, 16, 8, 120, 0]),
            ("4.3", "Нагрузочное тестирование (1000+ организаций)", "Подтверждение производительности и масштабируемости.", [0, 0, 0, 24, 8, 0, 48, 24]),
            ("4.4", "Тестирование безопасности и прав доступа", "Подтверждение безопасного разграничения доступа.", [0, 0, 8, 16, 8, 0, 40, 16]),
            ("4.5", "Тестирование офлайн-клиента (SMB, конфликты, backup)", "Подтверждение устойчивой офлайн-работы.", [0, 0, 0, 8, 0, 16, 40, 0]),
            ("4.6", "Исправление дефектов по результатам тестирования", "Стабилизированная версия перед ОПЭ.", [8, 0, 0, 48, 40, 24, 0, 0]),
        ],
    },
    {
        "id": "5",
        "name": "5. Подготовка к ОПЭ",
        "result": "Система развёрнута на стенде Заказчика, команды обучены, документация готова.",
        "works": [
            ("5.1", "Развёртывание на стенде Заказчика (Astra / Postgres Pro)", "Готовый стенд для опытной эксплуатации.", [8, 0, 8, 24, 0, 0, 16, 80]),
            ("5.2", "Подготовка эксплуатационной документации", "Комплект эксплуатационных и админских инструкций.", [8, 24, 0, 8, 8, 8, 8, 8]),
            ("5.3", "Обучение администраторов и методологов", "Обученная команда сопровождения Заказчика.", [8, 16, 0, 8, 8, 0, 0, 0]),
            ("5.4", "Подпись дистрибутивов desktop (корп. УЦ)", "Подписанные установщики для корпоративного контура.", [0, 0, 0, 0, 0, 16, 0, 24]),
            ("5.5", "Миграция production-данных из Access", "Подготовленная продуктивная база для запуска.", [0, 8, 0, 24, 0, 0, 16, 8]),
        ],
    },
    {
        "id": "6",
        "name": "6. Опытно-промышленная эксплуатация (ОПЭ)",
        "result": "Пилот в рабочих условиях подтверждён, замечания устранены, решение готово к ПРОД.",
        "works": [
            ("6.1", "Сопровождение пилотной группы (5–10 организаций, 1 квартал)", "Поддержка бизнеса в реальных сценариях.", [16, 16, 0, 40, 24, 16, 40, 16]),
            ("6.2", "Сбор замечаний и доработка по результатам ОПЭ", "Доработанная версия с учётом обратной связи.", [8, 8, 0, 40, 32, 8, 0, 0]),
            ("6.3", "Подтверждение критериев приёмки на реальных данных", "Акт закрытия ОПЭ и готовности к ПРОД.", [8, 16, 0, 8, 8, 0, 40, 0]),
        ],
    },
    {
        "id": "7",
        "name": "7. Передача в промышленную эксплуатацию",
        "result": "Система передана в ПРОД, команда Заказчика обучена, запуск завершён.",
        "works": [
            ("7.1", "Развёртывание production-контура (on-prem)", "Рабочий промышленный контур у Заказчика.", [8, 0, 8, 8, 0, 0, 8, 48]),
            ("7.2", "Обучение пользователей (исполнители)", "Подготовленные конечные пользователи.", [8, 16, 0, 0, 8, 8, 0, 0]),
            ("7.3", "Передача исходных кодов и документации", "Переданные материалы и регламенты сопровождения.", [8, 8, 0, 0, 0, 0, 0, 8]),
            ("7.4", "Подготовка документов для реестра российского ПО (при необходимости)", "Пакет документов для регуляторных процедур.", [8, 16, 8, 0, 0, 0, 8, 8]),
            ("7.5", "Гарантийное сопровождение (3 месяца после приёмки)", "Поддержка и устранение гарантийных замечаний.", [16, 8, 0, 24, 16, 8, 16, 8]),
        ],
    },
]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_range(ws, r1, r2, ncols):
    for r in range(r1, r2 + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Оценка по работам"
    ws["A1"] = "Оценка трудозатрат — ПК «ОКО» (коммерческое предложение)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")
    ws["A2"] = f"Дата: {date.today().strftime('%d.%m.%Y')} · Ставка: {RATE_PER_HOUR:,} руб/час".replace(",", " ")
    ws.merge_cells("A2:E2")

    headers = ["№", "Этап / работа", "Результат"] + ROLES + ["Итого, ч", "Ставка, руб/ч", "Стоимость, руб"]
    start_row = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, start_row, len(headers))

    row = start_row + 1
    stage_sums: list[tuple[str, int, int, str]] = []
    role_grand = [0] * len(ROLES)

    for stage in STAGES:
        ws.cell(row=row, column=1, value=stage["id"])
        ws.cell(row=row, column=2, value=stage["name"])
        ws.cell(row=row, column=3, value=stage["result"])
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = STAGE_FILL
            ws.cell(row=row, column=c).font = STAGE_FONT
        row += 1

        stage_roles = [0] * len(ROLES)
        for wid, work_name, work_result, hours in stage["works"]:
            ws.cell(row=row, column=1, value=wid)
            ws.cell(row=row, column=2, value=work_name)
            ws.cell(row=row, column=3, value=work_result)
            total_h = 0
            for i, h in enumerate(hours):
                if h:
                    ws.cell(row=row, column=4 + i, value=h)
                total_h += h
                stage_roles[i] += h
                role_grand[i] += h
            ws.cell(row=row, column=12, value=total_h)
            ws.cell(row=row, column=13, value=RATE_PER_HOUR)
            ws.cell(row=row, column=14, value=total_h * RATE_PER_HOUR)
            row += 1

        stage_hours = sum(stage_roles)
        stage_cost = stage_hours * RATE_PER_HOUR
        ws.cell(row=row, column=2, value=f"Итого по этапу {stage['id']}")
        ws.cell(row=row, column=3, value=stage["result"])
        ws.cell(row=row, column=2).font = TOTAL_FONT
        for i, v in enumerate(stage_roles):
            ws.cell(row=row, column=4 + i, value=v if v else "")
        ws.cell(row=row, column=12, value=stage_hours)
        ws.cell(row=row, column=13, value=RATE_PER_HOUR)
        ws.cell(row=row, column=14, value=stage_cost)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="EEF3F8")
        stage_sums.append((stage["id"], stage_hours, stage_cost, stage["name"]))
        row += 1
        row += 1

    grand_hours = sum(role_grand)
    grand_cost = grand_hours * RATE_PER_HOUR
    reserve_h = round(grand_hours * 0.15)
    reserve_cost = reserve_h * RATE_PER_HOUR

    ws.cell(row=row, column=2, value="ИТОГО ПО ПРОЕКТУ (без резерва)")
    ws.cell(row=row, column=2).font = TOTAL_FONT
    for i, v in enumerate(role_grand):
        ws.cell(row=row, column=4 + i, value=v)
    ws.cell(row=row, column=12, value=grand_hours)
    ws.cell(row=row, column=13, value=RATE_PER_HOUR)
    ws.cell(row=row, column=14, value=grand_cost)
    row += 1

    ws.cell(row=row, column=2, value="Резерв на риски 15%")
    ws.cell(row=row, column=12, value=reserve_h)
    ws.cell(row=row, column=13, value=RATE_PER_HOUR)
    ws.cell(row=row, column=14, value=reserve_cost)
    row += 1

    ws.cell(row=row, column=2, value="ИТОГО С РЕЗЕРВОМ")
    ws.cell(row=row, column=2).font = TOTAL_FONT
    ws.cell(row=row, column=12, value=grand_hours + reserve_h)
    ws.cell(row=row, column=13, value=RATE_PER_HOUR)
    ws.cell(row=row, column=14, value=grand_cost + reserve_cost)
    ws.cell(row=row, column=14).font = TOTAL_FONT

    style_range(ws, start_row, row, len(headers))
    for r in range(start_row + 1, row + 1):
        for c in range(4, 13):
            ws.cell(row=r, column=c).number_format = "#,##0"
        ws.cell(row=r, column=13).number_format = "#,##0"
        ws.cell(row=r, column=14).number_format = "#,##0"

    set_widths(ws, [5, 44, 36, 11, 12, 10, 12, 12, 12, 11, 11, 10, 12, 16])

    ws2 = wb.create_sheet("Сводка по этапам")
    ws2.append(["№", "Этап", "Часы", "Стоимость, руб", "% бюджета", "Результат"])
    style_header(ws2, 1, 6)
    for sid, sh, sc, sname in stage_sums:
        result = next(s["result"] for s in STAGES if s["id"] == sid)
        ws2.append([sid, sname, sh, sc, sc / grand_cost if grand_cost else 0, result])
    ws2.append(["", "ИТОГО (без резерва)", grand_hours, grand_cost, 1, ""])
    ws2.append(["", "Резерв 15%", reserve_h, reserve_cost, reserve_cost / grand_cost if grand_cost else 0, ""])
    ws2.append(["", "ИТОГО С РЕЗЕРВОМ", grand_hours + reserve_h, grand_cost + reserve_cost, "", ""])
    style_range(ws2, 2, ws2.max_row, 6)
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=3).number_format = "#,##0"
        ws2.cell(row=r, column=4).number_format = "#,##0"
        ws2.cell(row=r, column=5).number_format = "0%"
    set_widths(ws2, [4, 40, 10, 16, 10, 48])

    ws3 = wb.create_sheet("Сводка по ролям")
    ws3.append(["Роль", "Часы", "Стоимость, руб", "% от бюджета", "Чел.-мес. (160ч)"])
    style_header(ws3, 1, 5)
    for i, role in enumerate(ROLES):
        hrs = role_grand[i]
        cost = hrs * RATE_PER_HOUR
        ws3.append([role, hrs, cost, cost / grand_cost if grand_cost else 0, round(hrs / 160, 1)])
    ws3.append(["ИТОГО", grand_hours, grand_cost, 1, round(grand_hours / 160, 1)])
    style_range(ws3, 2, ws3.max_row, 5)
    for r in range(2, ws3.max_row + 1):
        ws3.cell(row=r, column=2).number_format = "#,##0"
        ws3.cell(row=r, column=3).number_format = "#,##0"
        ws3.cell(row=r, column=4).number_format = "0%"
    set_widths(ws3, [32, 10, 16, 12, 14])

    ws4 = wb.create_sheet("Допущения")
    ws4["A1"] = "Допущения и ограничения оценки"
    ws4["A1"].font = TITLE_FONT
    notes = [
        "1. Ставка принята: 4 000 руб/ч (единообразно для укрупненной оценки КП).",
        "2. Оценка включает полный цикл: обследование → проектирование → разработка → тестирование → ОПЭ → передача в ПРОД.",
        "3. Этапы 3–6 могут выполняться параллельно при укомплектованной команде.",
        "4. Резерв 15% заложен на изменения требований и инфраструктурные риски.",
        "5. SSO/LDAP, подпись desktop и реестр ПО зависят от инфраструктуры и регуляторных процедур Заказчика.",
        "6. Гарантийное сопровождение (3 месяца) включено.",
    ]
    for i, txt in enumerate(notes, 3):
        ws4.cell(row=i, column=1, value=txt)
    set_widths(ws4, [110])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(
        f"Saved: {OUT} | total={grand_hours} ч, reserve={reserve_h} ч, "
        f"budget={grand_cost + reserve_cost:,} руб".replace(",", " ")
    )


if __name__ == "__main__":
    main()
